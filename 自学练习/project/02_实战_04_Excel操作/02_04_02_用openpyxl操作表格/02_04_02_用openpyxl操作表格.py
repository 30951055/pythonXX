'''
openpyxl
    特点：
        不同于xlwt，它能操作xlsx
        不同于xlrd，xlwt，openpyxl的下标是从1开始的。
'''
from openpyxl.styles.fills import fills

'''
使用openpyxl写入数据
'''
import openpyxl
# 创建一个新的工作簿对象
wb = openpyxl.Workbook()
# 获取当前sheet
ws = wb.active
# 当前sheet的右边创建一个新的9月作业表sheet
ws2 = wb.create_sheet('9月作业表',1)
# 与xlrd，xlwt不同的是，openpyxl的下标是从1开始的。
ws2.cell(row=1,column=1,value='张三这个月没完成')
# 命名为text.xlsx并保存
wb.save('test.xlsx')
print('创建并写入数据后的结果是：',openpyxl.load_workbook('test.xlsx')['9月作业表']['A1'].value)


'''
使用openpyxl读取数据并修改
    相对于xlutils来说就方便太多了。
'''
wb = openpyxl.load_workbook('test.xlsx')
# 列出所有的sheet
print('test.xlsx文件中的所有sheet是：',wb.sheetnames)
# 选中'9月作业表'sheet
ws = wb['9月作业表']
# 更改sheet的颜色
ws.sheet_properties.tabColor = '99CCCC'
# 更改该sheet内cell的值
ws.cell(row=1,column=1,value='李四这个月没完成')
wb.save('test.xlsx')
print('修改后的结果是：',openpyxl.load_workbook('test.xlsx')['9月作业表']['A1'].value)


'''
修改样式
    Alignment 对齐
    merge_cells 合并单元格
    dimensions 距离
    italic 斜体
    vertical 垂直对齐方式 有bottom，center，top
    PatternFill 颜色填充
    wrap_text 自动换行

    设置单元格边框 border_style属性有dotted, hair, thin, mediumDashed, thick, double, mediumDashDotDot, dashDotDot, medium, slantDashDot, dashDot, dashed, mediumDashDot
    
    # 操作多列
        for column in ws['A:C']:
            for cell in column:
                print(cell.value)
    # 操作多行
        for row in ws['1:3']:
            for cell in row:
                print(cell.value)
    # 指定范围
        for row in ws['A1:C3']:
            for cell in row:
                print(cell.value)
    # 所有行
        for row in ws.iter_rows():
            for cell in row:
                print(cell.value)
    # 所有列
        for column in ws.iter_cols():
            for cell in column:
                print(cell.value)
'''
from openpyxl.styles import Font, colors, Alignment, PatternFill, Color, Protection, Border, Side
from openpyxl.comments import Comment

wb = openpyxl.Workbook()
ws = wb.active
rows = [
    ['序号','名字','年龄'],
    [1,'张三',25],
    [2,'李四',29],
    [3,'王五',33],
    [4,'赵六',19],
]
for row in rows:
    ws.append(row)
font = Font(name='微软雅黑',size=25,italic=True,color=colors.BLUE,bold=True)
ws['A1'].font = font
# 单元格颜色填充
ws['A3'].fill = PatternFill(fill_type='solid', start_color='FF0000')
# 设置单元格边框 border_style属性有dotted, hair, thin, mediumDashed, thick, double, mediumDashDotDot, dashDotDot, medium, slantDashDot, dashDot, dashed, mediumDashDot
ws['B5'].border = Border(left=Side(border_style='thick', color='FF0000'), right= Side(border_style='thin', color='FF0000'))
# 让A1单元格内容左对齐
ws['A1'].alignment = Alignment(horizontal='left')
# 给单元格加标注
ws['B3'].comment = Comment(text='这是一个批注',author='sunshine')
# 对指定区域进行水平垂直对齐
customStyle = Alignment
for i in ws['A1:C3']:
    # i是一个行对象，需要用j去遍历i才能得到每一行中的每个单元格
    for j in i:
        print('j的值是：',j)
        # 水平居左，垂直居上
        j.alignment = Alignment(horizontal='left',vertical='top')
# 对指定的行与列进行高度和宽度设置
ws.row_dimensions[5].height=40
ws.column_dimensions['A'].width=30
# cell单行合并
ws.merge_cells('A7:C7')
# 解除合并
# ws.unmerge_cells('A2:D2')
# 任意cell合并
ws.merge_cells('A9:C13')
ws['A9']='这是合并后的单元格123456这是合并后的单元格123456这是合并后的单元格123456这是合并后的单元格123456这是合并后的单元格123456这是合并后的单元格123456这是合并后的单元格123456这是合并后的单元格123456'
# 合并后的自动换行
ws['A9'].alignment = Alignment(wrap_text=True)

wb.save('test1.xlsx')



'''
使用openpyxl读取大批量的数据
    采用read_only和write_only
        read_only：False 全部读取  *默认
        read_only：True 按需读取
        write_only：False 全部写入  *默认
        write_only：True 按需写入
            1，只写模式下，不能修改已有工作表，只能create_sheet()方法去创建一个全新的。
            2，只写状态下，单元格想要具有样式或注释，一定要使用 openpyxl.cell.WriteOnlyCell()
            3，只写模式下，只能使用append()，而不能在任意位置用cell()或者iter_rows()
'''
# 读取超大量数据。
from openpyxl import load_workbook
wb = load_workbook(filename='big.xlsx',read_only=True)
ws = wb['big_sheet']
for row in ws.rows:
    for cell in row:
        print(cell.value)

# 写入超大量数据。
from openpyxl import Workbook
from openpyxl.cell import  WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import  Font
# write_only设置为True
wb = Workbook(write_only = True)
ws = wb.create_sheet()
cell = WriteOnlyCell(ws,value='只写状态下写入的内容')
cell.font = Font(name='微软雅黑',size=36)
# 插入Excel批注
cell.comment = Comment(text='这是一个批注',author='sunshine')
# 在只写工作簿中，只能使用append()
ws.append([cell,3.14,None])
wb.save('write_only_file.xlsx')