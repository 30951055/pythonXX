"""
关于yield
    只要在函数中有yield，那么它就是生成器函数而不是函数
        return  全部返回（自动返回全部结果）
        yield   间歇性返回（调用一次返回部分结果）
"""

import os, re
from openpyxl import load_workbook
import xlrd, time, sys



'''--------定义函数 START-----------'''
# 定义一个获取该地址下所有文件的函数
def file_names(dir):
    this_file = os.listdir(dir)
    for fname in this_file:
        f_dir = dir + '\\' + fname
        if os.path.isdir(f_dir):
            for fdir in file_names(f_dir):
                yield fdir
        else:
            yield f_dir

# 定义一个读取xlsx文件内容的函数
def handle_xlsx(abs_file):
    # 设置user_input, error为全局变量
    global user_input, error
    try:
        # 导入文件 如果该文件里面又公式，公式全都不计算(data_only=False(默认))；公式全都计算(data_only=True)
        wb = load_workbook(abs_file, data_only=True)
        # 遍历所有sheet名。
        for wsheet in wb.sheetnames:
            have_sheet = 0
            # 遍历所有行
            for thisrow in range(1, wb[wsheet].max_row + 1):
                # 遍历所有列
                for thiscolumn in range(1, wb[wsheet].max_column + 1):
                    # 判断用户输入的关键字是否存在每一行每一列中
                    if re.search(user_input, str(wb[wsheet].cell(thisrow, thiscolumn).value)):
                        # 计算存在的sheet数
                        if have_sheet == 0:
                            have_sheet = 1
                            yield f'\n----------{f}-{str(wsheet)}-----------\n'
                        # 遍历该sheet的所有列，并显示出所有行，列带有用户搜索关键字的内容
                        for thecolumn in range(1, wb[wsheet].max_column + 1):
                            t = str(wb[wsheet].cell(thisrow, thecolumn).value) + " "
                            yield t
                        yield "\n"
                        break
    except:
        error += 1
        yield "\n---------" + abs_file + "---------读取错误\n"

# 定义一个读取xls文件内容的函数    与读# 读取xlsx文件内容 逻辑大致相同。
def handle_xls(abs_file):
    global user_input, error
    try:
        wb = xlrd.open_workbook(abs_file)
        for wsheet in wb.sheet_names():
            have_sheet = 0
            # 遍历所有行 nrows：需要读取的行数(从文件头开始算起)
            for thisrow in range(0, wb.sheet_by_name(wsheet).nrows):
                # 遍历所有列 ncols：需要读取的列数(从文件头开始算起)
                for thiscolumn in range(0, wb.sheet_by_name(wsheet).ncols):
                    # 判断用户输入的关键字是否存在每一行每一列中
                    if re.search(user_input, str(wb.sheet_by_name(wsheet).cell(thisrow, thiscolumn).value)):
                        # 计算存在的sheet数
                        if have_sheet == 0:
                            have_sheet = 1
                            yield f'\n----------{f}-{str(wsheet)}-----------\n'
                        for thecolumn in range(0, wb.sheet_by_name(wsheet).ncols):
                            # 遍历该sheet的所有列，并显示出所有行，列带有用户搜索关键字的内容
                            t = str(wb.sheet_by_name(wsheet).cell(thisrow, thecolumn).value) + " "
                            yield t
                        yield "\n"
                        break
    except:
        error += 1
        yield "\n---------" + abs_file + "---------读取错误\n"

# 定义一个读取csv文件内容的函数
def handle_csv(abs_file):
    global user_input, error
    try:
        with open(abs_file) as f:
            have = 0
            for row in f:
                # 判断用户输入的关键字是否存在每一行中
                if re.search(user_input, row):
                    if have == 0: yield f'\n----------{abs_file}------------\n'
                    yield row
                    have = 1

    except:
        error += 1
        yield "\n---------" + abs_file + "---------读取错误\n"

# 将搜索结果整理到search_result.txt文件中并保存
def write_file(dir, content):
    with open(dir + "\\" + time.strftime("%Y-%m-%d", time.localtime())+'_search_result.txt', 'a', encoding='utf-8') as f:
    # with open(dir + "\\" + 'search_result.txt', 'a', encoding='utf-8') as f:
        f.write(content)

# 判断要搜索什么类型的文件
def file_handle(file, handle):
    global xlsx, xls, csv
    if file == "xlsx":
        xlsx += 1
    elif file == "xls":
        xls += 1
    elif file == "csv":
        csv += 1
    for result in handle(f):
        t = str(result)
        print(t)
        if save == "1": write_file(my_dir_path, t)
'''--------定义函数 END-----------'''


'''--------定义入口 START-----------'''
def u_input():
    # 设置save, user_input为全局变量
    global save, user_input
    save = input("1、程序递归搜索当前目录下所有xls/xlsx/csv文件！\n\
2、支持模糊搜索，支持正则表达式！\n\
3、如选择保存，保存的文件为当前目录下的search_result.txt \n\
---------------------------------------------------------\n\
第一步：请先选择是否需要保存搜索结果！按回车确定\n\
保存请输入： 1 \n\
不保存直接按回车\n")
    user_input = input("第二步：请输入需要搜索的内容或正则表达式！\n")
'''--------定义入口 END-----------'''




'''程序入口'''
zong = xlsx = xls = csv = error = 0
save = user_input = ''
u_input()

# 由于python脚本与打包后的exe文件获取的文件路径不一样，所以需要更改下面的代码
# 根据需要，取消下面两行其中一行代码的注释#
# my_dir_path = os.path.dirname(os.path.realpath(sys.executable)) # 打包成exe使用此句
# 获取当前路径
my_dir_path = os.path.dirname(os.path.abspath(__file__))  # 不打包，直接执行py脚本使用此句
file_top = "\n***************" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + \
           "***************\n当前路\
           径：" + my_dir_path + "\n                       \
搜索关键字：" + user_input + "\n\n"
for f in file_names(my_dir_path):
    zong += 1
    # 判断文件类型
    if re.search('\.xlsx$', f):
        # 如果该文件是xlsx文件，则读取xlsx文件内容
        file_handle("xlsx", handle_xlsx)
    elif re.search('\.xls$', f):
        # 如果该文件是xls文件，则读取xls文件内容
        file_handle("xls", handle_xls)
    elif re.search('\.csv$', f):
        # 如果该文件是csv文件，则读取csv文件内容
        file_handle("csv", handle_csv)
file_end = f'\n\n **********总文件{zong}个，其中xlsx文件{xlsx}个,xls文件{xls}\
个,csv文件{csv}个。错误文件{error}个。**********\n\n'

# 如果save为1，则执行写入到该文件的操作
if save == "1": write_file(my_dir_path, file_top)
if save == "1": write_file(my_dir_path, file_end)
input("按回车退出")
