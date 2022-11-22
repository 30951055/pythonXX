"""
关于yield
    只要在函数中有yield，那么它就是生成器函数而不是函数
        return  全部返回（自动返回全部结果）
        yield   间歇性返回（调用一次返回部分结果）
"""

import os, re

import numpy as np
import openpyxl
import pandas as pd
import xlrd, time, sys
# import xlrd, time

# import xlwt
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import label
from openpyxl.utils import get_column_letter
from pandas import ExcelWriter
# from goto import with_goto

wb = Workbook()
ws = wb.active


'''--------定义函数 START-----------'''
# 定义一个获取该地址下所有文件的函数
def file_names(dir):
    this_file = os.listdir(dir)
    for fname in this_file:
        f_dir = dir + '\\' + fname
        if os.path.isdir(f_dir):
            for fdir in file_names(f_dir):
                yield fdir
        else:
            yield f_dir

# 定义一个读取xlsx文件内容的函数
def handle_xlsx(abs_file):
    # 设置user_input, error为全局变量
    global user_input, error
    try:
        # 导入文件 如果该文件里面又公式，公式全都不计算(data_only=False(默认))；公式全都计算(data_only=True)
        wb = load_workbook(abs_file, data_only=True)
        # 遍历所有sheet名。
        for wsheet in wb.sheetnames:
            have_sheet = 0
            # 遍历所有行
            for thisrow in range(1, wb[wsheet].max_row + 1):
                # 遍历所有列
                for thiscolumn in range(1, wb[wsheet].max_column + 1):
                    # 判断用户输入的关键字是否存在每一行每一列中
                    if re.search(user_input, str(wb[wsheet].cell(thisrow, thiscolumn).value)):
                        # 计算存在的sheet数
                        if have_sheet == 0:
                            have_sheet = 1
                            yield f'\n{f}--\t{str(wsheet)}--\t'
                        # 遍历该sheet的所有列，并显示出所有行，列带有用户搜索关键字的内容
                        for thecolumn in range(1, wb[wsheet].max_column + 1):
                            t = str(wb[wsheet].cell(thisrow, thecolumn).value) + " "
                            yield t
                        yield "\n"
                        break
    except:
        error += 1
        yield "\n---------" + abs_file + "---------读取错误\n"

# 定义一个读取xls文件内容的函数    与读# 读取xlsx文件内容 逻辑大致相同。
def handle_xls(abs_file):
    global user_input, error
    try:
        wb = xlrd.open_workbook(abs_file)
        for wsheet in wb.sheet_names():
            have_sheet = 0
            # 遍历所有行 nrows：需要读取的行数(从文件头开始算起)
            for thisrow in range(0, wb.sheet_by_name(wsheet).nrows):
                # 遍历所有列 ncols：需要读取的列数(从文件头开始算起)
                for thiscolumn in range(0, wb.sheet_by_name(wsheet).ncols):
                    # 判断用户输入的关键字是否存在每一行每一列中
                    if re.search(user_input, str(wb.sheet_by_name(wsheet).cell(thisrow, thiscolumn).value)):
                        # 计算存在的sheet数
                        if have_sheet == 0:
                            have_sheet = 1
                            yield f'\n{f}--\t{str(wsheet)}--\t'
                        for thecolumn in range(0, wb.sheet_by_name(wsheet).ncols):
                            # 遍历该sheet的所有列，并显示出所有行，列带有用户搜索关键字的内容
                            t = str(wb.sheet_by_name(wsheet).cell(thisrow, thecolumn).value) + " "
                            yield t
                        yield "\n"
                        break
    except:
        error += 1
        yield "\n---------" + abs_file + "---------读取错误\n"

# 定义一个读取csv文件内容的函数
def handle_csv(abs_file):
    global user_input, error
    try:
        with open(abs_file) as f:
            have = 0
            for row in f:
                # 判断用户输入的关键字是否存在每一行中
                if re.search(user_input, row):
                    if have == 0: yield f'\n{abs_file}--\t\t'
                    yield row
                    have = 1

    except:
        error += 1
        yield "\n---------" + abs_file + "---------读取错误\n"

# 将搜索结果整理到txt文件中并保存
def write_txt_file(dir, content):
    with open(dir + "\\" + time.strftime("%Y-%m-%d", time.localtime())+'_search_result.txt', 'a', encoding='utf-8') as f:
        f.write(content)

# 将搜索结果整理到Excel文件中并保存
def write_excel_file(dir, contents):
    newTxtFile = dir + "\\" + time.strftime("%Y-%m-%d", time.localtime())+'_search_result.txt'
    with open(newTxtFile, 'a', encoding='utf-8') as f:
        f.write(contents)
        # txtToXls(newTxtFile,newTxtFile)




# 判断要搜索什么类型的文件
def file_handle(file, handle):
    global xlsx, xls, csv,fileBody
    if file == "xlsx":
        xlsx += 1
    elif file == "xls":
        xls += 1
    elif file == "csv":
        csv += 1
    for result in handle(f):
        fileBody = str(result)
        # 打印一下搜索到的内容
        # print(fileBody)
        # if save == "1": write_file(my_dir_path, fileBody)
        # 写入表体 把搜索到的fileBody内容写入到文件中。
        write_excel_file(my_dir_path, fileBody)

# txt转Excel
def txtToXls(dir):  #文本转换成xls的函数，filename 表示一个要被转换的txt文本，xlsname 表示转换后的文件名
    txtFile = dir + "\\" + time.strftime("%Y-%m-%d", time.localtime())+'_search_result.txt'
    with open(txtFile, "r", encoding="utf-8") as f:
        # 把逗号替换成统一的\t
        content = f.read()
        print(content)
        # 根据换行拆分内容
        lines = content.split("\n")
        # 提取标题
        # titles = lines[0].split("\t")
        titles = ['文件名','sheet名','目标内容']
        titles.insert(0, "")
        # 标题写入excel
        ws.append(titles)
        # 写入内容
        for i, line in enumerate(lines[1:]):
            if '总文件' in line or '搜索路径' in line or '搜索关键字' in line or '2022-'in line:
                continue
            item = line.split("\t")
            item.insert(0, i)

            ws.append(item)
    wb.save(txtFile.strip('.txt')+".xlsx")

    # 删除空行操作
    wb2 = openpyxl.load_workbook(txtFile.strip('.txt')+".xlsx")
    ws2 = wb2.active
    # シートの最終行を取得
    Sheet_Max = ws.max_row + 1

    # 最終行から逆ループ
    for i in reversed(range(1, Sheet_Max)):

        # A列が None だったら
        if ws2.cell(row=i, column=2).value == None:
            # 行削除
            ws2.delete_rows(i)

    wb2.save(txtFile.strip('.txt')+".xlsx")

    # 对表格进行重新排序
    df2 = pd.read_excel(txtFile.strip('.txt')+".xlsx",header=0)
    print(df2)
    df3 = df2.reset_index(drop=False)
    df3.index = df3.index+1
    # 删除B列
    df3.drop(df3.columns[1], axis=1, inplace=True)
    # df3.to_excel(txtFile.strip('.txt')+".xlsx")
    # 写入文件并自动调节列宽
    with pd.ExcelWriter(txtFile.strip('.txt')+".xlsx",engine='openpyxl') as writer:
        to_excel_auto_column_weight(df3, writer, f'Sheet1')
    delTxtFile(txtFile)


# 删除 txt文件
def delTxtFile(txt_file):
    os.remove(txt_file)



# 自动调节列宽
def to_excel_auto_column_weight(df: pd.DataFrame, writer: ExcelWriter, sheet_name):
    """DataFrame保存为excel并自动设置列宽"""
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    #  计算表头的字符宽度
    column_widths = (
        df.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
    )
    #  计算每列的最大字符宽度
    max_widths = (
        df.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
    )
    # 计算整体最大宽度
    widths = np.max([column_widths, max_widths], axis=0)
    # 设置列宽
    worksheet = writer.sheets[sheet_name]
    for i, width in enumerate(widths, 1):
        # openpyxl引擎设置字符宽度时会缩水0.5左右个字符，所以干脆+2使左右都空出一个字宽。
        worksheet.column_dimensions[get_column_letter(i)].width = width + 2

'''--------定义函数 END-----------'''



'''--------定义入口 START-----------'''
def u_input():
    # 设置save, user_input为全局变量
    global save, searchPath,user_input
#     temp = input("欢迎！本工具可以搜索任何目录下的xls/xlsx/csv文件！\n使用方法：\n1、输入要搜索的地址；\n\
# 2、支持模糊搜索，支持正则表达式！\n\
# 3、搜索到的结果将自动保存到当前的运行目录。 \n\
# ---------------------------------------------------------\n\
# 以上了解后，请按enter继续...\n")
    searchPath = input('\n"欢迎！本工具可以搜索任何目录下的xls/xlsx/csv文件！\n使用方法：\n1、输入要搜索的地址；\n\
2、支持模糊搜索，支持正则表达式！\n\
3、搜索到的结果将自动保存到当前的运行目录。 \n\
---------------------------------------------------------\n以上了解后，请输入你要搜索的路径：\n')
    user_input = input("请输入需要搜索的内容或正则表达式！\n")
'''--------定义入口 END-----------'''




'''程序入口'''
zong = xlsx = xls = csv = error = 0
save = user_input = ''
# 标识跳转并开始执行的地方
u_input()
# 由于python脚本与打包后的exe文件获取的文件路径不一样，所以需要更改下面的代码
# 根据需要，取消下面两行其中一行代码的注释#
# my_dir_path = os.path.dirname(os.path.realpath(sys.executable)) # 打包成exe使用此句
# 获取当前路径
my_dir_path = os.path.dirname(os.path.abspath(__file__))  # 不打包，直接执行py脚本使用此句
# 定义表头
file_top = "\n***************" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + \
           "***************\n搜索路径：" + searchPath + "\n搜索关键字：" + user_input + "\n\n"
# 写入表头
write_excel_file(my_dir_path, file_top)

# 定义表尾
file_end = f'\n\n**********总文件{zong}个，其中xlsx文件{xlsx}个,xls文件{xls}\
个,csv文件{csv}个。错误文件{error}个。**********\n\n'

# # 如果save为1，则执行写入到该文件的操作
# if save == "1": write_file(customDirPath, file_top)
# if save == "1": write_file(customDirPath, file_end)

# 写入表尾
write_excel_file(my_dir_path, file_end)


for f in file_names(searchPath):
    zong += 1
    # 判断文件类型
    if re.search('\.xlsx$', f):
        # 如果该文件是xlsx文件，则读取xlsx文件内容
        file_handle("xlsx", handle_xlsx)
    elif re.search('\.xls$', f):
        # 如果该文件是xls文件，则读取xls文件内容
        file_handle("xls", handle_xls)
    elif re.search('\.csv$', f):
        # 如果该文件是csv文件，则读取csv文件内容
        file_handle("csv", handle_csv)


txtToXls(my_dir_path)

input("\n-----------------------------\n已经输出xlsx文件到本文件的同目录下了。\n请按回车退出...")
# 在有跳转标识的地方开始执行
